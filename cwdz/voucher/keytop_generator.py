"""科拓手续费 / 提现凭证生成。"""

from __future__ import annotations

import logging
import shutil
from calendar import monthrange
from collections.abc import Callable
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from cwdz.processor.keytop_workbook import normalize_lot_name

logger = logging.getLogger(__name__)

KEYTOP_CLIENT_CODE = "110018"
KEYTOP_CLIENT_NAME = "厦门科拓通讯技术股份有限公司"
FEE_REFERENCE = "科拓手续费ktsxf"
WITHDRAW_REFERENCE = "科拓提现KTTX"
AUTO_WITHDRAW_DESC = "自动提现"
MANUAL_WITHDRAW_DESC = "手动提现"
WITHDRAW_FEE_DESC = "提现手续费"
WITHDRAW_DESCRIPTIONS = {AUTO_WITHDRAW_DESC, MANUAL_WITHDRAW_DESC}
FEE_AMOUNT_COLUMN = "手续费"
FEE_DATE_COLUMN = "交易日期"
SETTLEMENT_AMOUNT_COLUMN = "结算金额"
WITHDRAW_AMOUNT_COLUMN = "出账金额"
WITHDRAW_DATE_COLUMN = "入账日期"
VOUCHER_CREATOR = "0067754"
DATE_COLUMNS = ("记账日期", "业务日期", "辅助账业务日期")
DATE_NUMBER_FORMAT = "yyyy-mm-dd"


@dataclass
class ProjectMapping:
    sheet_name: str
    project_code: str
    project_name: str
    company_code: str
    bank_account_code: str
    bank_account_name: str


@dataclass
class KeytopVoucherResult:
    fee_path: Path
    withdrawal_path: Path
    fee_count: int
    withdrawal_count: int
    skipped: list[str]


def generate_keytop_vouchers(
    source_path: Path,
    account_mapping_path: Path,
    fee_template_path: Path,
    withdrawal_template_path: Path,
    output_dir: Path,
    *,
    period: str,
    on_progress: Callable[[str], None] | None = None,
) -> KeytopVoucherResult:
    """按项目汇总源数据，生成科拓手续费与提现凭证。"""
    source_path = Path(source_path)
    account_mapping_path = Path(account_mapping_path)
    fee_template_path = Path(fee_template_path)
    withdrawal_template_path = Path(withdrawal_template_path)
    output_dir = Path(output_dir)

    for path, label in [
        (source_path, "凭证源数据"),
        (account_mapping_path, "项目对应收款账户信息"),
        (fee_template_path, "手续费凭证模板"),
        (withdrawal_template_path, "提现凭证模板"),
    ]:
        if not path.is_file():
            raise FileNotFoundError(f"{label}文件不存在: {path}")

    output_dir.mkdir(parents=True, exist_ok=True)
    period_label, period_start, period_end, accounting_period = _parse_period(period)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    _report(f"读取源数据: {source_path.name}", on_progress)
    source_df = pd.read_excel(source_path)
    if "页签名称" not in source_df.columns:
        raise ValueError("源数据缺少「页签名称」列，请先完成第二步整理")

    _report(f"源数据共 {len(source_df)} 行", on_progress)
    _report_period_filter_stats(source_df, period_start, period_end, on_progress)

    fee_summary, fee_breakdown = _summarize_fee(source_df, period_start, period_end)
    withdraw_summary = _summarize_withdrawal(source_df, period_start, period_end)
    for name, extra in fee_breakdown.items():
        base, withdraw_fee = extra
        _report(
            f"  {name}: 手续费 {base:.2f} + 提现手续费 {withdraw_fee:.2f} = {fee_summary[name]:.2f}",
            on_progress,
        )
    _report(
        f"手续费项目 {len(fee_summary)} 个（{FEE_DATE_COLUMN} 在账期内，"
        f"合计 {FEE_AMOUNT_COLUMN} + {WITHDRAW_FEE_DESC} 的 |{SETTLEMENT_AMOUNT_COLUMN}|），"
        f"提现项目 {len(withdraw_summary)} 个（{WITHDRAW_DATE_COLUMN} 在账期内，合计 {WITHDRAW_AMOUNT_COLUMN}）",
        on_progress,
    )

    mappings, skipped = _load_project_mappings(
        account_mapping_path,
        set(fee_summary) | set(withdraw_summary),
    )
    for name in skipped:
        _report(f"  - 跳过未匹配项目: {name}", on_progress)

    fee_rows = _build_fee_rows(fee_summary, mappings, period_label, period_end, accounting_period)
    withdraw_rows = _build_withdraw_rows(
        withdraw_summary, mappings, period_label, period_end, accounting_period
    )

    fee_out = output_dir / f"凭证-科拓手续费_{period}_{stamp}.xlsx"
    withdraw_out = output_dir / f"凭证-科拓提现_{period}_{stamp}.xlsx"

    _report(f"生成手续费凭证 ({len(fee_rows) // 2} 个项目)…", on_progress)
    _write_voucher_file(fee_template_path, fee_out, fee_rows)

    _report(f"生成提现凭证 ({len(withdraw_rows) // 2} 个项目)…", on_progress)
    _write_voucher_file(withdrawal_template_path, withdraw_out, withdraw_rows)

    _report(f"手续费凭证: {fee_out.name}", on_progress)
    _report(f"提现凭证: {withdraw_out.name}", on_progress)
    return KeytopVoucherResult(
        fee_path=fee_out,
        withdrawal_path=withdraw_out,
        fee_count=len(fee_rows) // 2,
        withdrawal_count=len(withdraw_rows) // 2,
        skipped=skipped,
    )


def _summarize_fee(
    df: pd.DataFrame, period_start: date, period_end: date
) -> tuple[dict[str, float], dict[str, tuple[float, float]]]:
    """按页签名称汇总手续费（仅账期内数据，按交易日期判断）。

    - 常规行：合计「手续费」列
    - 结算说明为「提现手续费」：额外合计「结算金额」的绝对值
    - 账期外行不参与合计

    Returns:
        (项目合计, {项目: (手续费列合计, 提现手续费合计)})
    """
    if FEE_DATE_COLUMN not in df.columns:
        raise ValueError(f"源数据缺少「{FEE_DATE_COLUMN}」列")
    if FEE_AMOUNT_COLUMN not in df.columns:
        raise ValueError(f"源数据缺少「{FEE_AMOUNT_COLUMN}」列")
    if SETTLEMENT_AMOUNT_COLUMN not in df.columns:
        raise ValueError(f"源数据缺少「{SETTLEMENT_AMOUNT_COLUMN}」列")
    if "结算说明" not in df.columns:
        raise ValueError("源数据缺少「结算说明」列")

    working = df.copy()
    working[FEE_AMOUNT_COLUMN] = pd.to_numeric(working[FEE_AMOUNT_COLUMN], errors="coerce").fillna(0)
    working[SETTLEMENT_AMOUNT_COLUMN] = pd.to_numeric(
        working[SETTLEMENT_AMOUNT_COLUMN], errors="coerce"
    ).fillna(0)
    subset, excluded = _filter_in_period(working, FEE_DATE_COLUMN, period_start, period_end)
    if excluded:
        logger.info(
            "手续费汇总: 账期 %s ~ %s，按 %s 排除账期外 %d 行",
            period_start,
            period_end,
            FEE_DATE_COLUMN,
            excluded,
        )

    fee_totals = subset.groupby("页签名称", dropna=False)[FEE_AMOUNT_COLUMN].sum()
    withdraw_fee_rows = subset[subset["结算说明"].astype(str).str.strip() == WITHDRAW_FEE_DESC]
    withdraw_fee_totals = (
        withdraw_fee_rows.assign(
            _abs_settlement=withdraw_fee_rows[SETTLEMENT_AMOUNT_COLUMN].abs()
        )
        .groupby("页签名称", dropna=False)["_abs_settlement"]
        .sum()
    )

    all_names = set(fee_totals.index) | set(withdraw_fee_totals.index)
    result: dict[str, float] = {}
    breakdown: dict[str, tuple[float, float]] = {}
    for name in all_names:
        if pd.isna(name):
            continue
        base = round(float(fee_totals.get(name, 0)), 2)
        withdraw_fee = round(float(withdraw_fee_totals.get(name, 0)), 2)
        total = round(base + withdraw_fee, 2)
        if total > 0:
            key = str(name)
            result[key] = total
            if withdraw_fee > 0:
                breakdown[key] = (base, withdraw_fee)
    return result, breakdown


def _summarize_withdrawal(df: pd.DataFrame, period_start: date, period_end: date) -> dict[str, float]:
    """按页签名称汇总提现（仅账期内数据，按入账日期判断，限自动/手动提现）。"""
    if WITHDRAW_DATE_COLUMN not in df.columns:
        raise ValueError(f"源数据缺少「{WITHDRAW_DATE_COLUMN}」列")
    if WITHDRAW_AMOUNT_COLUMN not in df.columns:
        raise ValueError(f"源数据缺少「{WITHDRAW_AMOUNT_COLUMN}」列")
    if "结算说明" not in df.columns:
        return {}

    working = df.copy()
    working[WITHDRAW_AMOUNT_COLUMN] = pd.to_numeric(
        working[WITHDRAW_AMOUNT_COLUMN], errors="coerce"
    ).fillna(0)
    in_period, excluded = _filter_in_period(working, WITHDRAW_DATE_COLUMN, period_start, period_end)
    subset = in_period[
        in_period["结算说明"].astype(str).str.strip().isin(WITHDRAW_DESCRIPTIONS)
    ]
    if excluded:
        logger.info(
            "提现汇总: 账期 %s ~ %s，按 %s 排除账期外 %d 行",
            period_start,
            period_end,
            WITHDRAW_DATE_COLUMN,
            excluded,
        )
    grouped = subset.groupby("页签名称", dropna=False)[WITHDRAW_AMOUNT_COLUMN].sum()
    return {
        str(name): round(float(amount), 2)
        for name, amount in grouped.items()
        if pd.notna(name) and round(float(amount), 2) > 0
    }


def _load_project_mappings(
    mapping_path: Path,
    sheet_names: set[str],
) -> tuple[dict[str, ProjectMapping], list[str]]:
    projects = pd.read_excel(mapping_path, sheet_name="项目编码", dtype=str)
    accounts = pd.read_excel(mapping_path, sheet_name="项目对应收款账户", dtype=str)

    project_by_name: dict[str, tuple[str, str]] = {}
    for _, row in projects.iterrows():
        name = str(row.get("项目名称") or "").strip()
        code = str(row.get("项目编码") or "").strip()
        if name and code:
            project_by_name[normalize_lot_name(name)] = (code, name)

    account_by_name: dict[str, tuple[str, str, str, str]] = {}
    for _, row in accounts.iterrows():
        name = str(row.get("项目") or "").strip()
        if not name or name in account_by_name:
            continue
        account_by_name[name] = (
            str(row.get("公司编码") or "").strip(),
            str(row.get("收款账户编码") or "").strip(),
            str(row.get("账户") or "").strip(),
            name,
        )

    mappings: dict[str, ProjectMapping] = {}
    skipped: list[str] = []
    for sheet_name in sheet_names:
        account = account_by_name.get(sheet_name)
        project = project_by_name.get(normalize_lot_name(sheet_name))
        if not account or not project:
            skipped.append(sheet_name)
            continue
        company_code, bank_code, bank_name, project_name = account
        project_code, _ = project
        mappings[sheet_name] = ProjectMapping(
            sheet_name=sheet_name,
            project_code=project_code,
            project_name=project_name,
            company_code=company_code,
            bank_account_code=bank_code,
            bank_account_name=bank_name,
        )
    return mappings, skipped


def _build_fee_rows(
    summary: dict[str, float],
    mappings: dict[str, ProjectMapping],
    period_label: str,
    period_end: date,
    accounting_period: int,
) -> list[dict]:
    rows: list[dict] = []
    items = sorted(
        ((name, amount) for name, amount in summary.items() if name in mappings),
        key=lambda item: (mappings[item[0]].company_code, mappings[item[0]].project_name),
    )
    for sheet_name, amount in items:
        mapping = mappings[sheet_name]
        summary_text = f"应收{mapping.project_name}-{period_label}-临停-科拓代扣手续费"
        base = {
            "公司": mapping.company_code,
            "记账日期": period_end,
            "业务日期": period_end,
            "会计期间": accounting_period,
            "凭证号": "0001",
            "摘要": summary_text,
            "原币金额": amount,
            "辅助账摘要": summary_text,
            "参考信息": FEE_REFERENCE,
            "现金流量标记": 2,
            "辅助账业务日期": period_end,
            "制单人": VOUCHER_CREATOR,
        }
        rows.append(
            {
                **base,
                "分录号": 1,
                "科目": 64010237,
                "科目名称": "手续费",
                "方向": 1,
                "借方金额": amount,
                "贷方金额": None,
                "核算项目1": "项目",
                "编码1": mapping.project_code,
                "名称1": mapping.project_name,
            }
        )
        rows.append(
            {
                **base,
                "分录号": 2,
                "科目": 112204,
                "科目名称": "自主收费",
                "方向": 0,
                "借方金额": None,
                "贷方金额": amount,
                "核算项目1": "客户",
                "编码1": KEYTOP_CLIENT_CODE,
                "名称1": KEYTOP_CLIENT_NAME,
                "核算项目2": "项目",
                "编码2": mapping.project_code,
                "名称2": mapping.project_name,
            }
        )
    return rows


def _build_withdraw_rows(
    summary: dict[str, float],
    mappings: dict[str, ProjectMapping],
    period_label: str,
    period_end: date,
    accounting_period: int,
) -> list[dict]:
    rows: list[dict] = []
    items = sorted(
        ((name, amount) for name, amount in summary.items() if name in mappings),
        key=lambda item: (mappings[item[0]].company_code, mappings[item[0]].project_name),
    )
    for sheet_name, amount in items:
        mapping = mappings[sheet_name]
        summary_text = f"收到{mapping.project_name}-{period_label}-临停-科拓代扣"
        base = {
            "公司": mapping.company_code,
            "记账日期": period_end,
            "业务日期": period_end,
            "会计期间": accounting_period,
            "凭证号": "0001",
            "摘要": summary_text,
            "原币金额": amount,
            "辅助账摘要": summary_text,
            "参考信息": WITHDRAW_REFERENCE,
            "现金流量标记": 4,
            "辅助账业务日期": period_end,
            "制单人": VOUCHER_CREATOR,
        }
        rows.append(
            {
                **base,
                "分录号": 1,
                "科目": 100201,
                "科目名称": "活期",
                "方向": 1,
                "借方金额": amount,
                "贷方金额": None,
                "核算项目1": "银行账户",
                "编码1": mapping.bank_account_code,
                "名称1": mapping.bank_account_name,
            }
        )
        rows.append(
            {
                **base,
                "分录号": 2,
                "科目": 112204,
                "科目名称": "自主收费",
                "方向": 0,
                "借方金额": None,
                "贷方金额": amount,
                "核算项目1": "客户",
                "编码1": KEYTOP_CLIENT_CODE,
                "名称1": KEYTOP_CLIENT_NAME,
                "核算项目2": "项目",
                "编码2": mapping.project_code,
                "名称2": mapping.project_name,
            }
        )
    return rows


def _write_voucher_file(template_path: Path, output_path: Path, rows: list[dict]) -> None:
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    template_df = pd.read_excel(template_path)
    template_columns = list(template_df.columns)
    defaults = template_df.iloc[0].to_dict() if not template_df.empty else {}
    col_index = {name: idx + 1 for idx, name in enumerate(template_columns)}
    date_formats = {
        col_name: ws.cell(2, col_index[col_name]).number_format
        for col_name in DATE_COLUMNS
        if col_name in col_index and ws.max_row >= 2
    }

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for offset, row_data in enumerate(rows):
        excel_row = 2 + offset
        merged = {**defaults, **row_data}
        for col_name in template_columns:
            value = merged.get(col_name)
            if pd.isna(value):
                value = None
            if col_name in DATE_COLUMNS:
                value = _as_date(value)
            cell = ws.cell(row=excel_row, column=col_index[col_name], value=value)
            if col_name in DATE_COLUMNS and value is not None:
                cell.number_format = date_formats.get(col_name, DATE_NUMBER_FORMAT)

    wb.save(output_path)
    logger.info("凭证已保存: %s (%d 行)", output_path, len(rows))


def _parse_period(period: str) -> tuple[str, date, date, int]:
    text = period.strip()
    if not text:
        raise ValueError("请填写账期，例如 2026-05")

    normalized = text.replace("/", "-").replace(".", "-")
    parts = normalized.split("-")
    if len(parts) < 2:
        raise ValueError(f"账期格式无效: {period}")

    year = int(parts[0])
    month = int(parts[1])
    if month < 1 or month > 12:
        raise ValueError(f"账期月份无效: {period}")

    period_start = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    period_end = date(year, month, last_day)
    period_label = f"{year}.{month}月"
    return period_label, period_start, period_end, month


def _parse_flow_date(value) -> date | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        value = int(value)

    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    if " " in text:
        text = text.split(" ", 1)[0]
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) < 8:
        return None
    return date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))


def _in_period(value: date | None, period_start: date, period_end: date) -> bool:
    return value is not None and period_start <= value <= period_end


def _period_mask(series: pd.Series, period_start: date, period_end: date) -> pd.Series:
    return series.map(lambda v: _in_period(_parse_flow_date(v), period_start, period_end))


def _filter_in_period(
    df: pd.DataFrame,
    date_column: str,
    period_start: date,
    period_end: date,
) -> tuple[pd.DataFrame, int]:
    """按日期列过滤账期内数据，返回 (子集, 账期外排除行数)。"""
    mask = _period_mask(df[date_column], period_start, period_end)
    return df.loc[mask].copy(), int((~mask).sum())


def _report_period_filter_stats(
    df: pd.DataFrame,
    period_start: date,
    period_end: date,
    on_progress: Callable[[str], None] | None,
) -> None:
    period_label = f"{period_start} ~ {period_end}"
    if FEE_DATE_COLUMN in df.columns:
        fee_mask = _period_mask(df[FEE_DATE_COLUMN], period_start, period_end)
        _report(
            f"手续费：按 {FEE_DATE_COLUMN} 账期 {period_label}，"
            f"纳入 {int(fee_mask.sum())} 行，排除 {int((~fee_mask).sum())} 行",
            on_progress,
        )
    if WITHDRAW_DATE_COLUMN in df.columns:
        wd_mask = _period_mask(df[WITHDRAW_DATE_COLUMN], period_start, period_end)
        _report(
            f"提现：按 {WITHDRAW_DATE_COLUMN} 账期 {period_label}，"
            f"纳入 {int(wd_mask.sum())} 行，排除 {int((~wd_mask).sum())} 行",
            on_progress,
        )


def _as_date(value) -> date | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, pd.Timestamp):
        return value.date()
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _report(message: str, on_progress: Callable[[str], None] | None) -> None:
    logger.info(message)
    if on_progress:
        on_progress(message)
