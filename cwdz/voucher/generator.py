from __future__ import annotations

import logging
import shutil
from collections.abc import Callable
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# 凭证模板中数据起始行（1-based），按财务模板调整
DATA_START_ROW = 2


def generate_voucher(
    df: pd.DataFrame,
    template_path: Path,
    output_dir: Path,
    *,
    period: str | None = None,
    unsettled_start: date | None = None,
    unsettled_end: date | None = None,
    on_progress: Callable[[str], None] | None = None,
) -> Path:
    """基于 Excel 模板生成凭证文件。"""
    if not template_path.exists():
        raise FileNotFoundError(
            f"凭证模板不存在: {template_path}\n"
            "请将财务提供的 Excel 模板放到 voucher/template.xlsx"
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    period = period or datetime.now().strftime("%Y-%m")
    output_path = output_dir / f"凭证_{period}.xlsx"

    if unsettled_start and unsettled_end:
        _report(
            f"未到账区间: {unsettled_start.isoformat()} ~ {unsettled_end.isoformat()}",
            on_progress,
        )

    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active

    # TODO: 按财务模板列映射关系填写数据，并排除未到账区间
    for i, row in df.iterrows():
        excel_row = DATA_START_ROW + i
        ws.cell(row=excel_row, column=1, value=row.iloc[0] if len(row) > 0 else "")
        ws.cell(row=excel_row, column=2, value=row.iloc[1] if len(row) > 1 else "")

    wb.save(output_path)
    logger.info("凭证已生成: %s", output_path)
    return output_path


def _report(message: str, on_progress: Callable[[str], None] | None) -> None:
    logger.info(message)
    if on_progress:
        on_progress(message)
