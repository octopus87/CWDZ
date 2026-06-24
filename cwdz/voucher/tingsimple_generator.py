"""停简单代扣结算凭证生成（未到账 / 已到账手续费 / 银行存款）。"""

from __future__ import annotations

import logging
import shutil
from collections.abc import Callable
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from cwdz.processor.keytop_workbook import normalize_lot_name
from cwdz.voucher.settlement_dates import parse_period_range

logger = logging.getLogger(__name__)

TINGSIMPLE_CLIENT_CODE = "010114"
TINGSIMPLE_CLIENT_NAME = "北京停简单信息技术有限公司"
VOUCHER_CREATOR = "0067716"
VOUCHER_TYPE = "收款凭证"
CURRENCY = "BB01"
DATE_COLUMN = "核算开始日期"
COMPLETION_DATE_COLUMN = "完成日期"
FEE_COLUMN = "手续费金额"
ARRIVAL_COLUMN = "到账金额"
LOT_COLUMN = "停车场名称"
DATE_COLUMNS = ("记账日期", "业务日期", "辅助账业务日期")
DATE_NUMBER_FORMAT = "mm-dd-yy"
VOUCHER_SHEETS = ("凭证", "凭证 (2)")

REF_UNSETTLED = "TJDKQJS"
REF_FEE = "TJDJSSXF"
REF_BANK = "TJDJS"


@dataclass
class TingsimpleProjectMapping:
    lot_name: str
    project_name: str
    project_code: str
    company_code: str
    company_name: str
    bank_code: str
    bank_name: str


@dataclass
class TingsimpleVoucherResult:
    unsettled_path: Path
    fee_path: Path
    bank_path: Path
    unsettled_count: int
    fee_count: int
    bank_count: int
    skipped: list[str]


def generate_tingsimple_vouchers(
    source_path: Path,
    account_mapping_path: Path,
    unsettled_template_path: Path,
    fee_template_path: Path,
    bank_template_path: Path,
    output_dir: Path,
    *,
    period: str,
    on_progress: Callable[[str], None] | None = None,
) -> TingsimpleVoucherResult:
    """按项目汇总停简单对账数据，生成三类凭证。"""
    source_path = Path(source_path)
    account_mapping_path = Path(account_mapping_path)
    unsettled_template_path = Path(unsettled_template_path)
    fee_template_path = Path(fee_template_path)
    bank_template_path = Path(bank_template_path)
    output_dir = Path(output_dir)

    for path, label in [
        (source_path, "凭证源数据"),
        (account_mapping_path, "项目对应收款账户信息"),
        (unsettled_template_path, "代扣结算未到账凭证模板"),
        (fee_template_path, "代扣结算已到账凭证模板"),
        (bank_template_path, "银行存款凭证模板"),
    ]:
        if not path.is_file():
            raise FileNotFoundError(f"{label}文件不存在: {path}")

    output_dir.mkdir(parents=True, exist_ok=True)
    period_label, period_start, period_end, accounting_period = _parse_period(period)
    unsettled_posting_date = _first_day_next_month(period_end)
    unsettled_accounting_period = unsettled_posting_date.month
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ref_suffix = _reference_suffix(period)

    _report(f"读取源数据: {source_path.name}", on_progress)
    source_df = pd.read_excel(source_path)
    if LOT_COLUMN not in source_df.columns:
        raise ValueError(f"源数据缺少「{LOT_COLUMN}」列，请先完成第二步整理")
    if COMPLETION_DATE_COLUMN not in source_df.columns:
        raise ValueError(f"源数据缺少「{COMPLETION_DATE_COLUMN}」列")

    fee_summary = _summarize_fee(source_df, period_start, period_end)
    unsettled_summary = _summarize_arrival_unsettled(source_df, period_start, period_end)
    bank_summary = _summarize_arrival_settled(source_df, period_start, period_end)

    _report(
        f"账期 {period_label}，未到账：完成日期为空或不在本账期内",
        on_progress,
    )
    _report(
        f"汇总: 未到账 {len(unsettled_summary)} 项，"
        f"手续费 {len(fee_summary)} 项，银行存款 {len(bank_summary)} 项",
        on_progress,
    )

    lot_names = set(fee_summary) | set(unsettled_summary) | set(bank_summary)
    mappings, skipped = _load_project_mappings(account_mapping_path, lot_names)
    for name in skipped:
        _report(f"  - 跳过未匹配车场: {name}", on_progress)

    unsettled_rows = _build_bank_style_rows(
        unsettled_summary,
        mappings,
        period_label,
        unsettled_posting_date,
        unsettled_accounting_period,
        reference=f"{REF_UNSETTLED}-{ref_suffix}",
    )
    fee_rows = _build_fee_rows(
        fee_summary,
        mappings,
        period_label,
        period_end,
        accounting_period,
        reference=f"{REF_FEE}-{ref_suffix}",
    )
    bank_rows = _build_bank_style_rows(
        bank_summary,
        mappings,
        period_label,
        period_end,
        accounting_period,
        reference=f"{REF_BANK}-{ref_suffix}",
    )

    unsettled_out = output_dir / f"凭证-停简单代扣结算未到账_{period}_{stamp}.xlsx"
    fee_out = output_dir / f"凭证-停简单代扣结算已到账_{period}_{stamp}.xlsx"
    bank_out = output_dir / f"凭证-停简单银行存款_{period}_{stamp}.xlsx"

    _report(f"生成未到账凭证 ({len(unsettled_rows) // 2} 个项目)…", on_progress)
    _write_voucher_file(unsettled_template_path, unsettled_out, unsettled_rows)

    _report(f"生成已到账手续费凭证 ({len(fee_rows) // 2} 个项目)…", on_progress)
    _write_voucher_file(fee_template_path, fee_out, fee_rows)

    _report(f"生成银行存款凭证 ({len(bank_rows) // 2} 个项目)…", on_progress)
    _write_voucher_file(bank_template_path, bank_out, bank_rows)

    _report(f"未到账凭证: {unsettled_out.name}", on_progress)
    _report(f"手续费凭证: {fee_out.name}", on_progress)
    _report(f"银行存款凭证: {bank_out.name}", on_progress)

    return TingsimpleVoucherResult(
        unsettled_path=unsettled_out,
        fee_path=fee_out,
        bank_path=bank_out,
        unsettled_count=len(unsettled_rows) // 2,
        fee_count=len(fee_rows) // 2,
        bank_count=len(bank_rows) // 2,
        skipped=skipped,
    )


def _summarize_fee(df: pd.DataFrame, period_start: date, period_end: date) -> dict[str, float]:
    working = _prepare_source(df)
    in_period = _filter_by_date(working, period_start, period_end)
    grouped = in_period.groupby(LOT_COLUMN, dropna=False)[FEE_COLUMN].sum()
    return _positive_summary(grouped)


def _summarize_arrival_unsettled(
    df: pd.DataFrame,
    period_start: date,
    period_end: date,
) -> dict[str, float]:
    working = _prepare_source(df)
    in_period = _filter_by_date(working, period_start, period_end)
    if in_period.empty:
        return {}
    mask = _unsettled_mask(in_period, period_start, period_end)
    unsettled = in_period[mask.fillna(False)]
    grouped = unsettled.groupby(LOT_COLUMN, dropna=False)[ARRIVAL_COLUMN].sum()
    return _positive_summary(grouped)


def _summarize_arrival_settled(
    df: pd.DataFrame,
    period_start: date,
    period_end: date,
) -> dict[str, float]:
    working = _prepare_source(df)
    in_period = _filter_by_date(working, period_start, period_end)
    if in_period.empty:
        return {}
    mask = _unsettled_mask(in_period, period_start, period_end)
    settled = in_period[~mask.fillna(True)]
    grouped = settled.groupby(LOT_COLUMN, dropna=False)[ARRIVAL_COLUMN].sum()
    return _positive_summary(grouped)


def _unsettled_mask(df: pd.DataFrame, period_start: date, period_end: date) -> pd.Series:
    """完成日期为空或不在本账期内视为未到账。"""
    completion_dates = df[COMPLETION_DATE_COLUMN].map(_parse_completion_date)
    in_period = completion_dates.map(
        lambda value: value is not None and period_start <= value <= period_end
    )
    return ~in_period.fillna(True)


def _prepare_source(df: pd.DataFrame) -> pd.DataFrame:
    working = df.copy()
    if DATE_COLUMN not in working.columns:
        raise ValueError(f"源数据缺少「{DATE_COLUMN}」列")
    for col in (FEE_COLUMN, ARRIVAL_COLUMN):
        if col not in working.columns:
            raise ValueError(f"源数据缺少「{col}」列")
        working[col] = pd.to_numeric(working[col], errors="coerce").fillna(0)
    working[DATE_COLUMN] = pd.to_datetime(working[DATE_COLUMN], errors="coerce")
    if COMPLETION_DATE_COLUMN not in working.columns:
        raise ValueError(f"源数据缺少「{COMPLETION_DATE_COLUMN}」列")
    working[COMPLETION_DATE_COLUMN] = pd.to_datetime(
        working[COMPLETION_DATE_COLUMN], errors="coerce"
    )
    return working


def _filter_by_date(df: pd.DataFrame, start: date, end: date) -> pd.DataFrame:
    if df.empty:
        return df
    start_ts = pd.Timestamp(start)
    end_ts = pd.Timestamp(end)
    dates = df[DATE_COLUMN]
    mask = (dates >= start_ts) & (dates <= end_ts)
    return df[mask.fillna(False)]


def _positive_summary(grouped: pd.Series) -> dict[str, float]:
    result: dict[str, float] = {}
    for name, amount in grouped.items():
        if pd.isna(name):
            continue
        value = round(float(amount), 2)
        if value > 0:
            result[str(name).strip()] = value
    return result


def _load_project_mappings(
    mapping_path: Path,
    lot_names: set[str],
) -> tuple[dict[str, TingsimpleProjectMapping], list[str]]:
    df = pd.read_excel(mapping_path, sheet_name="项目对应收款账户信息", dtype=str)
    by_lot: dict[str, TingsimpleProjectMapping] = {}
    by_norm: dict[str, TingsimpleProjectMapping] = {}

    for _, row in df.iterrows():
        lot = str(row.get("车场名称") or "").strip()
        if not lot:
            continue
        mapping = TingsimpleProjectMapping(
            lot_name=lot,
            project_name=str(row.get("项目名称") or lot).strip(),
            project_code=str(row.get("项目编码") or "").strip(),
            company_code=str(row.get("公司编码") or "").strip(),
            company_name=str(row.get("公司") or "").strip(),
            bank_code=str(row.get("银行编码") or "").strip(),
            bank_name=str(row.get("银行") or "").strip(),
        )
        by_lot[lot] = mapping
        by_norm[normalize_lot_name(lot)] = mapping

    mappings: dict[str, TingsimpleProjectMapping] = {}
    skipped: list[str] = []
    for lot_name in lot_names:
        mapping = by_lot.get(lot_name) or by_norm.get(normalize_lot_name(lot_name))
        if mapping is None or not mapping.project_code or not mapping.bank_code:
            skipped.append(lot_name)
            continue
        mappings[lot_name] = mapping
    return mappings, skipped


def _build_bank_style_rows(
    summary: dict[str, float],
    mappings: dict[str, TingsimpleProjectMapping],
    period_label: str,
    posting_date: date,
    accounting_period: int,
    *,
    reference: str,
) -> list[dict]:
    rows: list[dict] = []
    items = sorted(
        ((name, amount) for name, amount in summary.items() if name in mappings),
        key=lambda item: (mappings[item[0]].company_code, mappings[item[0]].project_name),
    )
    for entry_no, (lot_name, amount) in enumerate(items, start=1):
        mapping = mappings[lot_name]
        summary_text = f"收到{mapping.project_name}{period_label}-停简单代扣"
        base = _base_row(
            mapping,
            summary_text,
            amount,
            posting_date,
            accounting_period,
            reference,
            entry_no,
        )
        rows.append(
            {
                **base,
                "科目": 100201,
                "科目名称": "活期",
                "方向": 1,
                "借方金额": amount,
                "贷方金额": None,
                "核算项目1": "银行账户",
                "编码1": mapping.bank_code,
                "名称1": mapping.bank_name,
                "核算项目2": None,
                "编码2": None,
                "名称2": None,
            }
        )
        rows.append(_credit_row(base, amount, mapping))
    return rows


def _build_fee_rows(
    summary: dict[str, float],
    mappings: dict[str, TingsimpleProjectMapping],
    period_label: str,
    posting_date: date,
    accounting_period: int,
    *,
    reference: str,
) -> list[dict]:
    rows: list[dict] = []
    items = sorted(
        ((name, amount) for name, amount in summary.items() if name in mappings),
        key=lambda item: (mappings[item[0]].company_code, mappings[item[0]].project_name),
    )
    for entry_no, (lot_name, amount) in enumerate(items, start=1):
        mapping = mappings[lot_name]
        summary_text = f"收到{mapping.project_name}{period_label}-停简单代扣"
        base = _base_row(
            mapping,
            summary_text,
            amount,
            posting_date,
            accounting_period,
            reference,
            entry_no,
        )
        rows.append(
            {
                **base,
                "科目": 64010237,
                "科目名称": "手续费",
                "方向": 1,
                "借方金额": amount,
                "贷方金额": None,
                "核算项目1": "项目",
                "编码1": mapping.project_code,
                "名称1": mapping.project_name,
                "核算项目2": None,
                "编码2": None,
                "名称2": None,
            }
        )
        rows.append(_credit_row(base, amount, mapping))
    return rows


def _base_row(
    mapping: TingsimpleProjectMapping,
    summary_text: str,
    amount: float,
    posting_date: date,
    accounting_period: int,
    reference: str,
    entry_no: int,
) -> dict:
    return {
        "公司": mapping.company_code,
        "记账日期": posting_date,
        "业务日期": posting_date,
        "会计期间": accounting_period,
        "凭证类型": VOUCHER_TYPE,
        "凭证号": "0037",
        "分录号": entry_no,
        "摘要": summary_text,
        "币种": CURRENCY,
        "汇率": 1,
        "原币金额": amount,
        "数量": 0,
        "单价": 0,
        "辅助账摘要": summary_text,
        "参考信息": reference,
        "现金流量标记": 4,
        "辅助账业务日期": posting_date,
        "制单人": VOUCHER_CREATOR,
    }


def _credit_row(base: dict, amount: float, mapping: TingsimpleProjectMapping) -> dict:
    return {
        **base,
        "科目": 112204,
        "科目名称": "自主收费",
        "方向": 0,
        "借方金额": None,
        "贷方金额": amount,
        "核算项目1": "客户",
        "编码1": TINGSIMPLE_CLIENT_CODE,
        "名称1": TINGSIMPLE_CLIENT_NAME,
        "核算项目2": "项目",
        "编码2": mapping.project_code,
        "名称2": mapping.project_name,
    }


def _write_voucher_file(template_path: Path, output_path: Path, rows: list[dict]) -> None:
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    for sheet_name in VOUCHER_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        template_df = pd.read_excel(template_path, sheet_name=sheet_name)
        template_columns = list(template_df.columns)
        defaults = template_df.iloc[0].to_dict() if not template_df.empty else {}
        col_index = {name: idx + 1 for idx, name in enumerate(template_columns)}
        date_formats = {
            col_name: ws.cell(2, col_index[col_name]).number_format
            for col_name in DATE_COLUMNS
            if col_name in col_index and ws.max_row >= 2
        }
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for offset, row_data in enumerate(rows):
            excel_row = 2 + offset
            merged = {**defaults, **row_data}
            for col_name in template_columns:
                value = merged.get(col_name)
                if pd.isna(value):
                    value = None
                if col_name in DATE_COLUMNS:
                    value = _as_date(value)
                cell = ws.cell(row=excel_row, column=col_index[col_name], value=value)
                if col_name in DATE_COLUMNS and value is not None:
                    cell.number_format = date_formats.get(col_name, DATE_NUMBER_FORMAT)
    wb.save(output_path)
    logger.info("凭证已保存: %s (%d 行)", output_path, len(rows))


def _parse_period(period: str) -> tuple[str, date, date, int]:
    period_start, period_end = parse_period_range(period)
    year = period_start.year
    month = period_start.month
    period_label = f"{year}.{month}月"
    return period_label, period_start, period_end, month


def _reference_suffix(period: str) -> str:
    period_start, _ = parse_period_range(period)
    return f"{period_start.year}.{period_start.month}"


def _first_day_next_month(period_end: date) -> date:
    if period_end.month == 12:
        return date(period_end.year + 1, 1, 1)
    return date(period_end.year, period_end.month + 1, 1)


def _parse_completion_date(value) -> date | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)) and value == 0:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    return _as_date(value)


def _as_date(value) -> date | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, pd.Timestamp):
        return value.date()
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def _report(message: str, on_progress: Callable[[str], None] | None) -> None:
    logger.info(message)
    if on_progress:
        on_progress(message)
