"""科拓批量任务 Excel 读写。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ACCOUNT_FLOW_HEADERS = [
    "入账日期",
    "交易日期",
    "结算金额",
    "手续费",
    "入账金额",
    "出账金额",
    "账户余额",
    "结算类型",
    "结算说明",
]

NO_PERMISSION_MSG = "当前账号无该车场权限，无法下载数据"
NO_PERMISSION_TAB_COLOR = "FF0000"


def normalize_lot_name(name: str) -> str:
    s = name.strip()
    for old, new in [(" ", ""), ("（", "("), ("）", ")"), ("＆", "&")]:
        s = s.replace(old, new)
    return s


def match_lot(sheet_name: str, lots: list[dict]) -> dict | None:
    """按页签名匹配账号可见车场。"""
    norm_sheet = normalize_lot_name(sheet_name)
    exact = [lot for lot in lots if normalize_lot_name(str(lot.get("name", ""))) == norm_sheet]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        return exact[0]

    partial: list[dict] = []
    for lot in lots:
        norm_lot = normalize_lot_name(str(lot.get("name", "")))
        if norm_sheet in norm_lot or norm_lot in norm_sheet:
            partial.append(lot)
    if len(partial) == 1:
        return partial[0]
    return None


def clear_sheet_data(ws: Worksheet) -> None:
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def write_sheet_rows(ws: Worksheet, rows: list[list]) -> None:
    clear_sheet_data(ws)
    ws.append(list(ACCOUNT_FLOW_HEADERS))
    for row in rows:
        ws.append(row)


def write_sheet_message(ws: Worksheet, message: str) -> None:
    clear_sheet_data(ws)
    ws.append(list(ACCOUNT_FLOW_HEADERS))
    ws.cell(row=2, column=1, value=message)


def mark_sheet_no_permission(ws: Worksheet) -> None:
    """无车场权限：写入提示并将页签标红。"""
    write_sheet_message(ws, NO_PERMISSION_MSG)
    ws.sheet_properties.tabColor = NO_PERMISSION_TAB_COLOR


def load_task_workbook(path: str | Path):
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"任务文件不存在: {workbook_path}")
    return load_workbook(workbook_path)


def save_task_workbook(workbook, path: str | Path) -> Path:
    workbook_path = Path(path)
    workbook.save(workbook_path)
    return workbook_path
